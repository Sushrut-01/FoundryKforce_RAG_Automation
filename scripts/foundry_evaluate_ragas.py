#!/usr/bin/env python3
"""
Enhanced RAG Evaluation using Foundry SDK and RAGAS Metrics.

Integrates:
- Azure AI Foundry SDK (AIProjectClient)
- RAGAS metrics via OpenAI API from Foundry
- Faithfulness: Response grounded in context
- Answer Relevance: Response relevant to query
- Context Recall: Relevant context retrieved
- Context Precision: All context is relevant
- Answer Correctness: Answer factually correct
"""

import json
import sys
import os
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass, asdict
import logging

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:  # noqa: F841
    Workbook = None  # type: ignore[assignment]
    PatternFill = None  # type: ignore[assignment]
    Font = None  # type: ignore[assignment]
    Alignment = None  # type: ignore[assignment]
    Border = None  # type: ignore[assignment]
    Side = None  # type: ignore[assignment]
    EXCEL_AVAILABLE = False

try:
    import numpy as np
except ImportError:  # noqa: F841
    np = None  # type: ignore[assignment]

try:
    from azure.ai.projects import AIProjectClient
    from azure.identity import DefaultAzureCredential
    FOUNDRY_AVAILABLE = True
except ImportError:  # noqa: F841
    AIProjectClient = None  # type: ignore[assignment]
    DefaultAzureCredential = None  # type: ignore[assignment]
    FOUNDRY_AVAILABLE = False

try:
    from ragas.metrics import (  # type: ignore[import]
        faithfulness,
        answer_relevance,
        context_recall,
        context_precision,
        answer_correctness
    )
    from ragas import evaluate  # type: ignore[import]
    from datasets import Dataset  # type: ignore[import]
    RAGAS_AVAILABLE = True
except ImportError:  # noqa: F841
    faithfulness = None  # type: ignore[assignment]
    answer_relevance = None  # type: ignore[assignment]
    context_recall = None  # type: ignore[assignment]
    context_precision = None  # type: ignore[assignment]
    answer_correctness = None  # type: ignore[assignment]
    evaluate = None  # type: ignore[assignment]
    Dataset = None  # type: ignore[assignment]
    RAGAS_AVAILABLE = False

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


@dataclass
class RAGSMetrics:
    """RAGAS evaluation metrics for RAG systems."""
    faithfulness: float = 0.0
    answer_relevance: float = 0.0
    context_recall: float = 0.0
    context_precision: float = 0.0
    answer_correctness: float = 0.0
    
    def to_dict(self) -> Dict[str, float]:
        """Convert metrics to dictionary."""
        return asdict(self)
    
    def average(self) -> float:
        """Get average of all metrics."""
        if np is None:
            values = [
                self.faithfulness, self.answer_relevance, self.context_recall,
                self.context_precision, self.answer_correctness
            ]
            return sum(values) / len(values)
        return float(np.mean([  # type: ignore[return-value]
            self.faithfulness, self.answer_relevance, self.context_recall,
            self.context_precision, self.answer_correctness
        ]))


class FoundryRAGEvaluator:
    """
    RAG Evaluator integrating Foundry SDK with RAGAS metrics.
    
    Foundry SDK Integration:
    - Uses AIProjectClient for Azure AI Foundry connection
    - Authenticates via DefaultAzureCredential
    - Uses OpenAI models from Foundry resource
    - Supports batch evaluation with RAGAS metrics
    """
    
    def __init__(
        self,
        foundry_endpoint: Optional[str] = None,
        foundry_project: Optional[str] = None,
        model_name: str = "gpt-4",
        use_ragas_real: bool = True
    ) -> None:
        """
        Initialize Foundry RAG Evaluator.
        
        Args:
            foundry_endpoint: Foundry resource endpoint URL
            foundry_project: Foundry project name
            model_name: LLM model to use (from Foundry)
            use_ragas_real: Use real RAGAS metrics (requires RAGAS + LLM)
        """
        self.model_name = model_name
        self.use_ragas_real = use_ragas_real
        self.results: List[Dict[str, Any]] = []
        self.openai_client: Optional[Any] = None
        
        # Initialize Foundry SDK if available
        if FOUNDRY_AVAILABLE:
            try:
                foundry_endpoint_val = foundry_endpoint or os.getenv("FOUNDRY_ENDPOINT")
                foundry_project_val = foundry_project or os.getenv("FOUNDRY_PROJECT")
                
                if foundry_endpoint_val and foundry_project_val:
                    logger.info(f"Initializing Foundry SDK with endpoint: {foundry_endpoint_val}")
                    
                    project_client = AIProjectClient(  # type: ignore[call-arg]
                        endpoint=foundry_endpoint_val,
                        credential=DefaultAzureCredential()  # type: ignore[arg-type]
                    )
                    
                    # Get OpenAI client from Foundry
                    self.openai_client = project_client.get_openai_client()
                    logger.info("Successfully initialized Foundry OpenAI client")
                else:
                    logger.warning("FOUNDRY_ENDPOINT or FOUNDRY_PROJECT not set")
            except Exception as err:  # noqa: BLE001
                logger.warning(f"Could not initialize Foundry SDK: {err}")
        else:
            if use_ragas_real:
                logger.warning("Foundry SDK not available. Install: pip install azure-ai-projects")
        
        # Check RAGAS availability
        if use_ragas_real and not RAGAS_AVAILABLE:
            logger.warning("RAGAS not available. Install: pip install ragas datasets")
            self.use_ragas_real = False
        
        logger.info(f"Initialized FoundryRAGEvaluator (RAGAS: {self.use_ragas_real})")
    
    def validate_format(self, test_case: Dict[str, Any]) -> Tuple[bool, List[str]]:
        """Validate Foundry SDK format requirements."""
        errors: List[str] = []
        
        if "query" not in test_case:
            errors.append("Missing 'query' field")
        elif not isinstance(test_case["query"], str) or not test_case["query"].strip():
            errors.append("'query' must be non-empty string")
        
        if "response" not in test_case:
            errors.append("Missing 'response' field")
        elif not isinstance(test_case["response"], str):
            errors.append("'response' must be string")
        
        if "context" not in test_case:
            errors.append("Missing 'context' field")
        elif not isinstance(test_case["context"], (list, str)):
            errors.append("'context' must be list or string")
        
        return len(errors) == 0, errors
    
    def normalize_context(self, context: Any) -> List[str]:
        """Normalize context to list of strings."""
        if isinstance(context, str):
            return [context] if context.strip() else []
        elif isinstance(context, list):
            return [str(c) for c in context if c]
        return []
    
    def compute_ragas_real_scores(
        self, query: str, response: str, context: List[str]
    ) -> RAGSMetrics:
        """
        Compute RAGAS metrics using real LLM evaluation.
        
        Uses Foundry OpenAI client to score:
        - Faithfulness
        - Answer Relevance
        - Context Recall
        - Context Precision
        - Answer Correctness
        """
        try:
            if not RAGAS_AVAILABLE:
                logger.error("RAGAS library not available")
                return RAGSMetrics()
            
            # Create dataset for RAGAS evaluation
            eval_data = {
                "question": [query],
                "answer": [response],
                "contexts": [[" ".join(context)] if context else [""]],
                "ground_truth": [response]
            }
            
            dataset = Dataset.from_dict(eval_data)  # type: ignore[attr-defined]
            
            # Select evaluation metrics
            metrics = [
                faithfulness,
                answer_relevance,
                context_recall,
                context_precision,
                answer_correctness
            ]
            
            # Evaluate using Foundry's OpenAI client
            logger.info(f"Evaluating with RAGAS using {self.model_name}...")
            results = evaluate(  # type: ignore[call-arg]
                dataset=dataset,
                metrics=metrics,
                llm=self.openai_client
            )
            
            # Extract scores
            return RAGSMetrics(
                faithfulness=float(results["faithfulness"][0]),  # type: ignore[index]
                answer_relevance=float(results["answer_relevance"][0]),  # type: ignore[index]
                context_recall=float(results["context_recall"][0]),  # type: ignore[index]
                context_precision=float(results["context_precision"][0]),  # type: ignore[index]
                answer_correctness=float(results["answer_correctness"][0])  # type: ignore[index]
            )
            
        except Exception as err:  # noqa: BLE001
            logger.error(f"Error computing real RAGAS metrics: {err}")
            return RAGSMetrics()
    
    def compute_ragas_mock_scores(
        self, query: str, response: str, context: List[str]
    ) -> RAGSMetrics:
        """
        Compute mock RAGAS scores (no LLM calls).
        
        Fallback for testing without real evaluation.
        """
        try:
            query_words = set(query.lower().split())
            response_words = set(response.lower().split())
            context_text = " ".join(context).lower() if context else ""
            context_words = set(context_text.split())
            
            # Faithfulness: Response content grounded in context
            faithfulness_score = (
                len(response_words & context_words) / max(len(response_words), 1)
                if context else 0.5
            )
            
            # Answer Relevance: Query keywords in response
            answer_relevance_score = (
                len(query_words & response_words) / max(len(query_words), 1)
            )
            
            # Context Recall: Query keywords in context
            context_recall_score = (
                len(query_words & context_words) / max(len(query_words), 1)
                if context else 0.0
            )
            
            # Context Precision
            context_precision_score = 0.85 if context else 0.0
            
            # Answer Correctness
            answer_correctness_score = 0.75 if len(response) > 20 else 0.5
            
            return RAGSMetrics(
                faithfulness=min(max(faithfulness_score, 0.0), 1.0),
                answer_relevance=min(max(answer_relevance_score, 0.0), 1.0),
                context_recall=min(max(context_recall_score, 0.0), 1.0),
                context_precision=min(max(context_precision_score, 0.0), 1.0),
                answer_correctness=min(max(answer_correctness_score, 0.0), 1.0)
            )
        except Exception as err:  # noqa: BLE001
            logger.error(f"Error computing mock scores: {err}")
            return RAGSMetrics()
    
    def compute_ragas_scores(
        self, query: str, response: str, context: List[str]
    ) -> RAGSMetrics:
        """Compute RAGAS scores using real or mock metrics."""
        if self.use_ragas_real:
            return self.compute_ragas_real_scores(query, response, context)
        else:
            return self.compute_ragas_mock_scores(query, response, context)
    
    def evaluate_case(self, case_id: int, test_case: Dict[str, Any]) -> Dict[str, Any]:
        """Evaluate a single test case."""
        format_valid, format_errors = self.validate_format(test_case)
        
        query = test_case.get("query", "")
        response = test_case.get("response", "")
        context = self.normalize_context(test_case.get("context", []))
        
        if not format_valid:
            logger.warning(f"Format error in case {case_id}: {format_errors}")
            return {
                "case_id": case_id,
                "query": query,
                "response": response,
                "context": context,
                "metrics": RAGSMetrics().to_dict(),
                "format_valid": False,
                "status": "FORMAT_ERROR",
                "errors": format_errors
            }
        
        metrics = self.compute_ragas_scores(query, response, context)
        avg_score = metrics.average()
        
        if avg_score >= 0.8:
            status = "PASS"
        elif avg_score >= 0.6:
            status = "WARNING"
        else:
            status = "FAIL"
        
        result: Dict[str, Any] = {
            "case_id": case_id,
            "query": query[:60],
            "response": response[:60],
            "context_count": len(context),
            "metrics": metrics.to_dict(),
            "average_score": float(avg_score),
            "format_valid": True,
            "status": status
        }
        
        self.results.append(result)
        return result
    
    def evaluate_batch(self, test_cases: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Evaluate batch of test cases."""
        logger.info(f"Starting batch evaluation of {len(test_cases)} cases")
        
        results = [
            self.evaluate_case(i + 1, case)
            for i, case in enumerate(test_cases)
        ]
        
        logger.info(f"Completed batch evaluation - {len(results)} cases")
        return results
    
    def get_summary(self) -> Dict[str, Any]:
        """Get evaluation summary."""
        if not self.results:
            return {"message": "No results yet", "total_cases": 0}
        
        valid_results = [r for r in self.results if r['format_valid']]
        status_counts: Dict[str, int] = {}
        
        for r in self.results:
            status = r['status']
            status_counts[status] = status_counts.get(status, 0) + 1
        
        avg_metrics: Dict[str, float] = {}
        if valid_results:
            for metric in valid_results[0]['metrics'].keys():
                scores = [r['metrics'][metric] for r in valid_results]
                if np:
                    avg_metrics[metric] = float(np.mean(scores))  # type: ignore[arg-type]
                else:
                    avg_metrics[metric] = sum(scores) / len(scores)
        
        return {
            "total_cases": len(self.results),
            "valid_cases": len(valid_results),
            "status_breakdown": status_counts,
            "average_metrics": avg_metrics,
            "model": self.model_name,
            "using_real_ragas": self.use_ragas_real,
            "results": self.results
        }
    
    def print_report(self) -> None:
        """Print formatted evaluation report."""
        summary = self.get_summary()
        
        ragas_mode = "REAL RAGAS (LLM-based)" if summary['using_real_ragas'] else "MOCK RAGAS (heuristic)"
        
        print(f"\n{'='*80}")
        print("FOUNDRY RAG EVALUATION REPORT")
        print(f"{'='*80}\n")
        
        print(f"Model: {summary.get('model', 'Unknown')}")
        print(f"Mode: {ragas_mode}")
        print(f"Total Test Cases: {summary['total_cases']}")
        print(f"Valid Cases: {summary['valid_cases']}")
        print(f"\nStatus Breakdown:")
        for status, count in summary['status_breakdown'].items():
            pct = (count / summary['total_cases']) * 100 if summary['total_cases'] > 0 else 0
            print(f"  {status}: {count}/{summary['total_cases']} ({pct:.1f}%)")
        
        print(f"\nAverage RAGAS Metrics (0.0-1.0 scale):")
        for metric, score in summary.get('average_metrics', {}).items():
            bar = "#" * int(score * 10)
            print(f"  {metric:20s}: {score:.4f} [{bar:<10}]")
        
        print(f"\n{'='*80}")
        print("RAGAS Metrics Definitions:")
        print(f"  - Faithfulness: Response grounded in retrieved context")
        print(f"  - Answer Relevance: Response directly addresses query")
        print(f"  - Context Recall: Relevant context documents retrieved")
        print(f"  - Context Precision: All retrieved context is relevant")
        print(f"  - Answer Correctness: Answer factually accurate")
        print(f"{'='*80}\n")
    
    def export_to_excel(self, output_path: Path) -> bool:
        """
        Export evaluation results to Excel file.
        
        Creates formatted spreadsheet with:
        - Summary sheet with metrics overview
        - Detailed sheet with per-case results
        """
        if not EXCEL_AVAILABLE:
            logger.warning("openpyxl not available. Install: pip install openpyxl")
            return False
        
        try:
            summary = self.get_summary()
            wb = Workbook()  # type: ignore[attr-defined]
            
            # Remove default sheet and create new ones
            wb.remove(wb.active)  # type: ignore[union-attr]
            
            # ===== SUMMARY SHEET =====
            ws_summary = wb.create_sheet("Summary", 0)  # type: ignore[attr-defined]
            
            # Header styling
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # type: ignore[arg-type]
            header_font = Font(bold=True, color="FFFFFF", size=12)  # type: ignore[arg-type]
            
            # Summary Section
            ws_summary["A1"] = "FOUNDRY RAG EVALUATION SUMMARY"  # type: ignore[index]
            ws_summary["A1"].font = header_font  # type: ignore[index]
            ws_summary["A1"].fill = header_fill  # type: ignore[index]
            ws_summary.merge_cells("A1:B1")  # type: ignore[attr-defined]
            
            row = 3
            ws_summary[f"A{row}"] = "Model"  # type: ignore[index]
            ws_summary[f"B{row}"] = summary.get("model", "Unknown")  # type: ignore[index]
            row += 1
            
            ws_summary[f"A{row}"] = "Evaluation Mode"  # type: ignore[index]
            mode_text = "REAL RAGAS (LLM-based)" if summary["using_real_ragas"] else "MOCK RAGAS (heuristic)"
            ws_summary[f"B{row}"] = mode_text  # type: ignore[index]
            row += 1
            
            ws_summary[f"A{row}"] = "Total Test Cases"  # type: ignore[index]
            ws_summary[f"B{row}"] = summary["total_cases"]  # type: ignore[index]
            row += 1
            
            ws_summary[f"A{row}"] = "Valid Cases"  # type: ignore[index]
            ws_summary[f"B{row}"] = summary["valid_cases"]  # type: ignore[index]
            row += 2
            
            # Status Breakdown
            ws_summary[f"A{row}"] = "Status Breakdown"  # type: ignore[index]
            ws_summary[f"A{row}"].font = Font(bold=True, size=11)  # type: ignore[index]
            row += 1
            
            for status, count in summary["status_breakdown"].items():
                ws_summary[f"A{row}"] = status  # type: ignore[index]
                ws_summary[f"B{row}"] = count  # type: ignore[index]
                row += 1
            
            row += 1
            
            # Average Metrics
            ws_summary[f"A{row}"] = "Average RAGAS Metrics"  # type: ignore[index]
            ws_summary[f"A{row}"].font = Font(bold=True, size=11)  # type: ignore[index]
            row += 1
            
            for metric, score in summary.get("average_metrics", {}).items():
                ws_summary[f"A{row}"] = metric  # type: ignore[index]
                ws_summary[f"B{row}"] = round(score, 4)  # type: ignore[index]
                row += 1
            
            # Set column widths
            ws_summary.column_dimensions["A"].width = 25  # type: ignore[attr-defined]
            ws_summary.column_dimensions["B"].width = 20  # type: ignore[attr-defined]
            
            # ===== DETAILED RESULTS SHEET =====
            ws_results = wb.create_sheet("Results", 1)  # type: ignore[attr-defined]
            
            # Header row
            headers = [
                "Case ID", "Query", "Response", "Context Count",
                "Faithfulness", "Answer Relevance", "Context Recall",
                "Context Precision", "Answer Correctness", "Average Score", "Status"
            ]
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws_results.cell(row=1, column=col_idx)  # type: ignore[attr-defined]
                cell.value = header  # type: ignore[attr-defined]
                cell.font = header_font  # type: ignore[attr-defined]
                cell.fill = header_fill  # type: ignore[attr-defined]
                cell.alignment = Alignment(horizontal="center", vertical="center")  # type: ignore[attr-defined]
            
            # Data rows
            for row_idx, result in enumerate(summary["results"], 2):
                ws_results.cell(row=row_idx, column=1).value = result["case_id"]  # type: ignore[attr-defined]
                ws_results.cell(row=row_idx, column=2).value = result["query"]  # type: ignore[attr-defined]
                ws_results.cell(row=row_idx, column=3).value = result["response"]  # type: ignore[attr-defined]
                ws_results.cell(row=row_idx, column=4).value = result["context_count"]  # type: ignore[attr-defined]
                
                metrics = result["metrics"]
                ws_results.cell(row=row_idx, column=5).value = round(metrics["faithfulness"], 4)  # type: ignore[attr-defined]
                ws_results.cell(row=row_idx, column=6).value = round(metrics["answer_relevance"], 4)  # type: ignore[attr-defined]
                ws_results.cell(row=row_idx, column=7).value = round(metrics["context_recall"], 4)  # type: ignore[attr-defined]
                ws_results.cell(row=row_idx, column=8).value = round(metrics["context_precision"], 4)  # type: ignore[attr-defined]
                ws_results.cell(row=row_idx, column=9).value = round(metrics["answer_correctness"], 4)  # type: ignore[attr-defined]
                ws_results.cell(row=row_idx, column=10).value = round(result["average_score"], 4)  # type: ignore[attr-defined]
                ws_results.cell(row=row_idx, column=11).value = result["status"]  # type: ignore[attr-defined]
            
            # Format metric columns as decimal
            for row_idx in range(2, len(summary["results"]) + 2):
                for col_idx in range(5, 11):
                    ws_results.cell(row=row_idx, column=col_idx).number_format = "0.0000"  # type: ignore[attr-defined]
            
            # Set column widths
            col_widths = {
                "A": 10, "B": 30, "C": 30, "D": 15,
                "E": 15, "F": 18, "G": 15, "H": 18, "I": 18, "J": 15, "K": 10
            }
            for col, width in col_widths.items():
                ws_results.column_dimensions[col].width = width  # type: ignore[attr-defined]
            
            # Save workbook
            wb.save(output_path)  # type: ignore[union-attr]
            logger.info(f"Excel report saved to {output_path}")
            return True
            
        except Exception as err:  # noqa: BLE001
            logger.error(f"Error exporting to Excel: {err}")
            return False


def main() -> int:
    """Main execution with Foundry SDK and RAGAS integration."""
    base_path = Path(__file__).parent.parent
    input_override = os.getenv("TEST_CASES_PATH")
    default_kb_path = base_path / "data" / "processed" / "test_cases_with_kb.json"
    default_formatted_path = base_path / "data" / "processed" / "test_cases_formatted.json"
    test_cases_path = Path(input_override) if input_override else (
        default_kb_path if default_kb_path.exists() else default_formatted_path
    )
    
    if not test_cases_path.exists():
        logger.error(f"Test cases not found: {test_cases_path}")
        return 1
    
    try:
        with open(test_cases_path, 'r') as f:
            data = json.load(f)
    except json.JSONDecodeError as err:
        logger.error(f"Invalid JSON: {err}")
        return 1
    except FileNotFoundError as err:
        logger.error(f"File not found: {err}")
        return 1
    
    if isinstance(data, list):
        test_cases = data
    elif isinstance(data, dict) and "test_cases" in data:
        test_cases = data["test_cases"]
    else:
        test_cases = [data]
    
    logger.info(f"Loaded {len(test_cases)} test cases")
    
    # Determine if we should use real RAGAS
    use_real_ragas = RAGAS_AVAILABLE and FOUNDRY_AVAILABLE
    if not use_real_ragas:
        logger.warning("Using MOCK RAGAS (heuristic scoring)")
        if not RAGAS_AVAILABLE:
            logger.info("To enable real RAGAS: pip install ragas datasets")
        if not FOUNDRY_AVAILABLE:
            logger.info("To enable Foundry SDK: pip install azure-ai-projects")
    
    # Initialize evaluator
    evaluator = FoundryRAGEvaluator(
        foundry_endpoint=os.getenv("FOUNDRY_ENDPOINT"),
        foundry_project=os.getenv("FOUNDRY_PROJECT"),
        model_name="gpt-4",
        use_ragas_real=use_real_ragas
    )
    
    # Run batch evaluation
    evaluator.evaluate_batch(test_cases)
    evaluator.print_report()
    
    # Save results
    summary = evaluator.get_summary()
    results_path = base_path / "artifacts" / "latest" / "ragas_evaluation.json"
    results_path.parent.mkdir(parents=True, exist_ok=True)
    
    with open(results_path, 'w') as f:
        json.dump(summary, f, indent=2)
    
    logger.info(f"Results saved to {results_path}")
    
    # Export to Excel
    excel_path = base_path / "artifacts" / "latest" / "ragas_evaluation.xlsx"
    evaluator.export_to_excel(excel_path)
    
    pass_count = summary['status_breakdown'].get('PASS', 0)
    fail_count = summary['status_breakdown'].get('FAIL', 0)
    
    print(f"\nSummary: {pass_count} PASS, {fail_count} FAIL")
    return 0  # Return success if evaluation completed (pass/fail tracking is just for analysis)


if __name__ == "__main__":
    sys.exit(main())
